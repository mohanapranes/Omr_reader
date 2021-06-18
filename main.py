import uuid
import os
import shutil
from openpyxl import Workbook

from fastapi import FastAPI, File, UploadFile, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from typing import List
from app import omr
from app.email import email
from pydantic import BaseModel , EmailStr

from fastapi.responses import HTMLResponse

from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

origins = [
    "http://localhost:80","*"
]


app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)



async def save_file(path,file):
        tmpImg1 = open(path,'wb')
        tmpImg1.write(file.file.read())
        tmpImg1.close()

def delete_user_dir(user_id):
    shutil.rmtree(user_id) 

def create_sheet(user_id, student_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "STUDENT ID"
    sheet["B1"] = "TEST ID"
    index = 0
    for char in range(67,130):
        Char = chr(char)
        sheet[f"{Char}1"] = "QUESTION {index}"
        index += 1
        if index == 19:
            break
    for row, mark_sheet in enumerate(student_list):
        sheet[f"A{row+2}"] = str(mark_sheet["roll_num"])
        sheet[f"B{row+2}"] = mark_sheet["test_id"]
        index = 0 
        for char in range(67,130):
            Char = chr(char)
            sheet[f"{Char}{row+2}"] = mark_sheet["mark_sheet"][index]
            index += 1
            if index == 19:
                break 
    workbook.save(filename=f"{user_id}/mark_sheet.xlsx")

def email_sheet(path, receiver_email):
    message = email.attachSheet(path)
    email.send(receiver_email,message)

def read_data_vertical(arr):
    data = 0
    for y in range(len(arr[0])):
        for x in range(len(arr)):
            if arr[x][y] == 1:
                if x+1 != 1:
                    data = (data*10)+ (x+1)
                else:
                    data = (data*10)+ 0
    return data


def vald_horizontal(sheet_arr, answer_arr):
    for index in range(0,len(answer_arr)):
        if sheet_arr[index] != answer_arr[index]:
            return False
    return True

def calculate_mark(sheet_dict, answer_dict):
    mark_sheet_dict={}
    mark_sheet = []
    
    for key in range(1,21):
        temp = vald_horizontal(sheet_dict[key], answer_dict[key])
        mark_sheet.append(temp)
    
    mark_sheet_dict["roll_num"]   = read_data_vertical(sheet_dict["roll_number"])
    mark_sheet_dict["test_id"]    = read_data_vertical(sheet_dict["test_id"])
    mark_sheet_dict["mark_sheet"] = mark_sheet    
    return mark_sheet_dict



def fetch_dict(path):
    return omr.OMR.main(path)

html_content = """
<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>OMR-grader</title>
  <meta name="description" content="simple CV example created with HTML and CSS">
  <meta name="author" content="Fly Nerd">
  <link rel="icon" href="static/favicon.ico">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <link rel="stylesheet" href="https://use.fontawesome.com/releases/v5.0.13/css/all.css" integrity="sha384-DNOHZ68U8hZfKXOrtjWvjxusGo9WQnrNx2sqG0tfsghAvtVlRW3tvkXWZh58N9jp" crossorigin="anonymous">
  <link rel="stylesheet" href= "static/style.css" rel="stylesheet">
</head>

<body>
  <header>
    <div>
      <img src="static/ICON.png" />
    </div>
    <h1>SCRAP</h1>
    <section>
      <p>Hello! Here is the new OMR grader which reduce your time and work!!Very Happy to introduce this technology to the world!! </p>
      <a href="https://www.facebook.com/bitsathyindia/" target="_blank">
        <i class="fab fa-facebook-f"></i>
      </a>
      <a href="https://twitter.com/bitsathyindia?lang=en" target="_blank">
        <i class="fab fa-twitter"></i>
      </a>
      
      
      <a href="https://www.linkedin.com/school/bannari-amman-institute-of-technology-bannariamman-educational-trust-/about/" target="_blank">
        <i class="fab fa-linkedin-in"></i>
      </a>
    </section>
  </header>
  <main>
    <section>
      <h3>OMR  Evaluator.</h3>
      
    <form action="http://127.0.0.1:8000/validate/" method="POST" id="form" enctype="multipart/form-data">
    
        <strong>Your email</strong><br><br>
        <input type="email" name="client_mail" >
        <script>var input = document.createElement("input");
            input.type = "submit";
            input.value="UPLOAD";
            form.appendChild(input);
        </script>
        <br>
        <br>
        <strong>Answer Key</strong><br><br>
        <input type="file"  name="answerKey"><br><br>
    
        <strong>Answer sheets </strong>
        <input type="button" value="+" kabi@gmq.com onclick = "add()"><br><br>     
  </form>
    <script>
      function add(){
      var input1 = document.createElement('INPUT');
      input1.setAttribute("type","file");
      input1.setAttribute("name","files");
      form.appendChild(input1);
      var breaker = document.createElement('br');
      form.appendChild(breaker);
      console.log("added");
      var breaker = document.createElement('br');
      form.appendChild(breaker);
      console.log("added");
      }
     </script>
       
    </section>
    <section>
      <h3>SCRAP</h3>
      <div class='skills'>
        <div class='column'>
          <h4>TEAMMATES:</h4>
          <ul>
            <li>Mohanapraneswaran M</li>
            <li>Sanjay kumar S</li>
            <li>Kabilan M</li>
            <li>Nishanth</li>
          </ul>
        </div>
        <div class='column'>
          <h4>Department</h4>
          <ul>
            <li>Information Technology</li>
            <li>Information Technology</li>
            <li>Electronics and communication Engineering</li>
            <li>Electronics and communication Engineering</li>
          </ul>
        </div>
        
    </section>

    <section>
      <h3>Education</h3>
      
     <article>
        <div class='school'>
          <span>2020</span> <strong>Bannari amman Institute of Technology.</strong>
        </div>
        <div>
          Location: Alathucombai,Sathyamangalam,638 401.
        </div>
      </article>
    </section>
    <section>
      <h3>Technical Info</h3>
      <article>
        <div>
          <p><strong>Tech Used:</strong> HTML / CSS / Computer Vision / Fast API</p>
          <ul class="job-description">
            <li>User Friendly Interface</li>
            <li>Cost efficient</li>
           
          </ul>
        </div>
      </article>
    </section>
  </main>
  <footer>
    <p>Created by: <a href="#">@SMKN</a>  / <a href="#">LinkedIn</a> / 2020 </p>
  </footer>
</body>
</html>

"""

@app.get("/", response_class=HTMLResponse)
async def renderHtml():
    return html_content


@app.post("/validate/")
async def image_fun(client_mail:EmailStr = File(...),answerKey: UploadFile = File(...),files: List[UploadFile] = File(...)): 

    user_id = str(uuid.uuid1().int)
    file_count = len(files)
    file_names = list()
    path = user_id+ '/0'
    answer_file = path
    os.mkdir(user_id)
    await save_file(path, answerKey)
    
    for index, file in enumerate(files):
        path = user_id+ '/' +str(index+1)
        file_names.append(path)
        await save_file(path, file)

    answer_sheet = fetch_dict(user_id + '/0')
    student_list =[]
    for index in range(1, file_count+1):
        sheet_dict = fetch_dict(f"{user_id}/{index}")
        student_list.append(calculate_mark(sheet_dict,answer_sheet))
    create_sheet(user_id, student_list)
    email_sheet(f"{user_id}/mark_sheet.xlsx", client_mail)
    delete_user_dir(user_id)
    return {"status":"Excel sheet has been mailed"}